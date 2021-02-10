# -*- coding: utf-8 -*-

import os
import time
import warnings
from threading import Thread

import openpyxl
from openpyxl import load_workbook

from utils import time_track


class ParserKTO(Thread):
    # В значениях словарей уазаны номера столбцов, для считывания значений (начиная с 0 для екселя)
    search_scheme_title_new = {
        'Номер объекта:': 1,
        'Наименование:': 1,
        'Адрес:': 1,
        'Дата проведения работ:': 1,
        'Исполнитель 1:': 2
    }
    search_scheme_title_old = {
        'Код ЕРП:': 1,
        'Наименование:': 1,
        'Адрес:': 1,
        'Дата проведения работ:': 1,
        'Испольнитель:': 1
    }
    search_scheme_EPU = {
        'Тип ЭПУ:': [6],
        'Тип нагрузки': [6],
        'Выходной ток (общий), А': [6],
        'Состояние ЭПУ': [5],
        'Количество групп АКБ:': [5],
        'Количество АКБ в группе:': [5],
        'Тип АКБ:': [5, 6, 7, 8],
        'Всего АКБ в данном ЭПУ': [5],
        'Сумма номинальных емкостей АКБ, Ач:': [5],
        'Вывод': [1, 4],
        'Время автономной работы ориентировочно': [3],
        'Время автономной работы ориентировочно:': [3],
        'Расчетное время на АКБ:': [3]
    }
    search_scheme_EPU_old = {
        'Тип системы электропитания:': 3,
        'Выходное напряжение (общее): ': 7,
        'Результаты проверки выпрямительных модулей:': 5,
        'Тип аккумуляторных батарей:': 3,
        'Количество аккумуляторных батарей:': 3,
        'Заключение: ': 2,
        'Замена батареи / элемента ': 2
    }
    table_header = [
        'Номер объекта',
        'Наименование',
        'Адрес',
        'Дата проведения работ',
        'Исполнитель 1',
        'Тип ЭПУ',
        'Тип нагрузки',
        'Выходной ток (общий), А',
        'Состояние ЭПУ',
        'Количество групп АКБ',
        'Количество АКБ в группе',
        'Тип АКБ (группа 1)',
        'Тип АКБ (группа 2)',
        'Тип АКБ (группа 3)',
        'Тип АКБ (группа 4)',
        'Всего АКБ в данном ЭПУ',
        'Сумма номинальных емкостей АКБ, А/ч',
        'Вывод',
        'Количество (Вывод)',
        'Время автономной работы (ориентировочно)',
        'Расчетное время на АКБ',
        'Ссылка на отчет'
    ]

    def __init__(self, task, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.list_files = []
        self.directory = task[1]
        self.file_to_create = {
            'Name': task[0],
            'sheet.title': "Данные КТО",
            'end_row_number': 1
        }
        self.read_file = None
        self.my_wb = None
        self.my_sheet = None
        self.percentage_of_completion = 0
        self.count_process = 0
        self.total_files = None
        self.start_parsing = False

    def run(self):
        self.list_files = self._get_list_of_file(self.directory)
        self.start_parsing = True
        if not self.list_files:
            print(f'В директории "{self.directory}" отчетов КТО в формате exel не найдено.')
            self.percentage_of_completion = float(100)
            return
        self.total_files = len(self.list_files)
        self._create_file()

        for self.read_file in self.list_files:
            try:
                wb = load_workbook(self.read_file, data_only=True)
                if self.read_file.endswith('.xlsm'):
                    self._checking_new_report(wb)
                elif self.read_file.endswith('.xlsx'):
                    self._checking_old_report(wb)
                elif self.read_file.endswith('.XLSX'):
                    self._checking_old_report(wb)
            except Exception as exc:
                log = open((self.file_to_create['Name'][:-4]+'_log.txt'), 'a', encoding='UTF-8')
                error_str = f'{self.read_file} --> {type(exc)} {exc} \n'
                log.write(error_str)
                log.close()
            finally:
                self.count_process += 1
                self.percentage_of_completion = round(((self.count_process * 100) / self.total_files), 1)
        try:
            self.my_wb.save(self.file_to_create['Name'])
        except Exception as exc:
            print(f"Ошибка записи в файл {self.file_to_create['Name']}, закройте файл и перезапустите программу.")
            log = open((self.file_to_create['Name'][:-4]+'_log.txt'), 'a', encoding='UTF-8')
            error_str = f"{self.read_file} --> Ошибка записи в файл {self.file_to_create['Name']}. {type(exc)} {exc} \n"
            log.write(error_str)
            log.close()

    def _checking_new_report(self, wb):
        title_data_for_write = []
        epu_data_for_write = []
        sh_names = wb.sheetnames
        # Читаю титульный лист
        ws = wb[sh_names[0]]
        for index, row in enumerate(ws.iter_rows()):
            val = row[0].value
            if val is None:
                continue
            for key, shift in ParserKTO.search_scheme_title_new.items():
                if val == key:
                    # print(f'{row[0].value} {row[shift].value}')
                    title_data_for_write.append(row[shift].value)
                    break
        # Читаю отчет ЭПУ (все листы)
        for sh_name in sh_names:
            if sh_name[:3] == 'ЭПУ':
                ws = wb[sh_name]
                if ws.sheet_state == 'hidden':
                    continue
                # print(f'==========={sh_name}============')
                for index, row in enumerate(ws.iter_rows()):
                    val = row[0].value
                    if val is None:
                        continue
                    for key, shifts in ParserKTO.search_scheme_EPU.items():
                        if val == key:
                            for shift in shifts:
                                # print(f'{row[0].value} {row[shift].value}')
                                epu_data_for_write.append(row[shift].value)
                            break
                epu_data_for_write.append(self.read_file)
                write_row = title_data_for_write + epu_data_for_write
                self._write_row_to_file(data_for_write=write_row)
                epu_data_for_write = []

    def _checking_old_report(self, wb):
        title_data_for_write = []
        epu_data_for_write = []
        temp_data = []
        sh_names = wb.sheetnames
        # Читаем титульный лист
        ws = wb[sh_names[0]]
        # дополнительная проверка, что это именно отчет КТО ("А13" содерижит "Владелец объекта:")
        if ws['A13'].value == "Владелец объекта:":
            for index, row in enumerate(ws.iter_rows()):
                val = row[0].value
                if val is None:
                    continue
                for key, shift in ParserKTO.search_scheme_title_old.items():
                    if val == key:
                        title_data_for_write.append(row[shift].value)
                        break
            # Читаем отчет ЭПУ (все листы)
            for sh_name in sh_names:
                if sh_name[:14] == 'Электропитание':
                    ws = wb[sh_name]
                    # скрытые листы не читаем
                    if ws.sheet_state == 'hidden':
                        continue
                    for index, row in enumerate(ws.iter_rows()):
                        val = row[0].value
                        if val is None:
                            continue
                        for key, shift in ParserKTO.search_scheme_EPU_old.items():
                            if val == key:
                                if row[shift].value is None:
                                    temp_data.append('')
                                else:
                                    temp_data.append(row[shift].value)
                                break
                    # структура данных в title_data_for_write для старого и нового отчета одинаковая
                    # структуру данных для epu_data_for_write нужно преобразовать к формату нового отчета
                    epu_data_for_write.append(temp_data[0])  # Тип системы электропитания
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(temp_data[1])  # Выходной ток (в строке: Выходное напряжение (общее):)
                    epu_data_for_write.append(temp_data[2])  # Результаты проверки выпрямительных модулей
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(temp_data[3])  # Тип аккумуляторных батарей
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(temp_data[4])  # Количество аккумуляторных батарей
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(f'{temp_data[5]}. {temp_data[6]}')  # Заключение + Замена батареи/элемента
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(None)
                    epu_data_for_write.append(self.read_file)
                    temp_data = []
                    write_row = title_data_for_write + epu_data_for_write
                    self._write_row_to_file(data_for_write=write_row)
                    epu_data_for_write = []

    def _create_file(self):
        self.my_wb = openpyxl.Workbook()
        self.my_sheet = self.my_wb.active
        self.my_sheet.title = self.file_to_create["sheet.title"]
        self._write_row_to_file(data_for_write=ParserKTO.table_header)

    def _write_row_to_file(self, data_for_write):
        for column_number in range(len(data_for_write)):
            write_cell = self.my_sheet.cell(row=self.file_to_create['end_row_number'], column=column_number + 1)
            write_cell.value = data_for_write[column_number]
        self.file_to_create['end_row_number'] += 1

    def _get_list_of_file(self, dirpath):
        list_of_files = []
        normal_dir = os.path.normpath(dirpath)
        print('Ищем файлы в указанных директориях. Это может занять несколько минут.')
        try:
            for dirpath, _, filenames in os.walk(normal_dir):
                for filename in filenames:
                    link = os.path.join(dirpath, filename)
                    if filename[:26] == 'Форма устранения замечаний':
                        continue
                    elif filename[:32] == 'Копия Форма устранения замечаний':
                        continue
                    elif filename[:2] == '~$':
                        continue
                    size = os.path.getsize(link)
                    if size > 204800:
                        continue
                    elif filename.endswith('.xlsm'):
                        list_of_files.append(link)
                    elif filename.endswith('.xlsx'):
                        list_of_files.append(link)
                    elif filename.endswith('.XLSX'):
                        list_of_files.append(link)
        except Exception as exc:
            log = open((self.file_to_create['Name'][:-4]+'_log.txt'), 'a', encoding='UTF-8')
            error_str = f'Ошибка поиска файлов в: {dirpath} --> {type(exc)} {exc} \n'
            log.write(error_str)
            log.close()

        return list_of_files


@time_track
def main():
    warnings.filterwarnings("ignore", category=UserWarning)
    # warnings.filterwarnings("ignore", module=openpyxl)

    tasks = []
    try:
        # читаем файл с заданиями формата: имя_нового_файла.xlsx,адрес файла
        with open('task.txt', 'r', encoding='UTF8') as file:
            for line in file:
                tuple_task = line.split(sep=',')
                # строки начинающиеся на "#" пропускаем
                if tuple_task[0][:1] == '#':
                    continue
                tuple_task[0] = f'./Отчеты/{tuple_task[0]}'
                if tuple_task[1].endswith('\n'):
                    tuple_task[1] = tuple_task[1][:-1]
                tasks.append(tuple_task)
    except Exception as exc:
        print(f'Error read file task.txt: {exc}')
        quit()

    parsers = [ParserKTO(task=task) for task in tasks]

    count_thread = len(parsers)

    for parser in parsers:
        parser.start()

    # ждем окончания поиска файлов для парсинга, чтобы начать выводить прогресс парсинга.
    process_flag = False
    while not process_flag:
        time.sleep(1)
        for parser in parsers:
            if parser.start_parsing:
                process_flag = True

    print('Начинаем парсинг найденных файлов.')
    while process_flag:
        sum_proc = 0
        for parser in parsers:
            print(f"{parser.file_to_create['Name'][2:]:<40} --> {parser.percentage_of_completion:>5} %")
            sum_proc += parser.percentage_of_completion
            if parser.percentage_of_completion == float(100):
                process_flag = False
            else:
                process_flag = True
        print('====================================================')
        print(f'Парсинг выпонен на {round(sum_proc / count_thread, 1)} %')
        print('====================================================')
        time.sleep(15)

    for parser in parsers:
        parser.join()


if __name__ == '__main__':
    main()
