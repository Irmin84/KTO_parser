import pyexcel

array_dictionary = {'Sheet 1': [
                                ['ID', 'AGE', 'SCORE'],
                                [1, 22, 5],
                                [2, 15, 6],
                                [3, 28, 9]
                                ],
                    'Sheet 2': [
                                ['X', 'Y', 'Z'],
                                [1, 2, 3],
                                [4, 5, 6],
                                [7, 8, 9]
                                ],
                    'Sheet 3': [
                                ['M', 'N', 'O', 'P'],
                                [10, 11, 12, 13],
                                [14, 15, 16, 17],
                                [18, 19, 20, 21]
                                ]}

# Save the data to a file                        
pyexcel.save_book_as(bookdict=array_dictionary, dest_file_name="2d_array_data.xls")

# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
#https://codecamp.ru/blog/python-excel-tutorial/

# import csv
from itertools import islice
from pprint import pprint
# Import `pandas`
import pandas as pd

from openpyxl import load_workbook
# Import relevant modules from `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string





# invent = []
#
# # with open('test.xlsx', 'r', encoding='UTF8', newline='') as xlsx_file:
# #     xlsx_file = csv.reader(xlsx_file)
# #     for row in xlsx_file:
# #         invent.append(row)
#
# # Load in the workbook
wb = load_workbook('./test.xlsx')
# # Помните, что вы можете изменить рабочий каталог с помощью os.chdir()
sh_names = wb.sheetnames
sheet = wb[sh_names[1]]

# Convert Sheet to DataFrame
df = pd.DataFrame(sheet.values)
# Put the sheet values in `data`
data = sheet.values

# Indicate the columns in the sheet values
cols = next(data)[1:]

# Convert your data to a list
data = list(data)

# Read in the data at index 0 for the indices
idx = [r[0] for r in data]

# Slice the data at index 1
data = (islice(r, 1, None) for r in data)

# Make your DataFrame
df = pd.DataFrame(data, index=idx, columns=cols)

# pprint(data)
print(df)



# sh = wb.sheetnames
# print(sh)
# # sh = wb.get_sheet_names()
# # # print(wb.active)
# #
# #
# #
# # Get a sheet by name
# sheet = wb[sh[1]]
# # sheet = wb.get_sheet_by_name(sh[1])
# print(sheet['A1'].value)
# # Print the sheet title
# # sheet.title
#
# # Get currently active sheet
# anotherSheet = wb.active
#
# # Check `anotherSheet`
# anotherSheet

# c = sheet['B2']
# print(c.row)
# print(c.column)
# print(c.coordinate)
#
# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)
#
# # Return 'A'
# print(get_column_letter(1))
#
# # Return '1'
# print(column_index_from_string('A'))
#
# # Print row per row
# for cellObj in sheet['A1':'C3']:
#       for cell in cellObj:
#               print(cell.coordinate, cell.value)
#       print('--- END ---')
#
# # Retrieve the maximum amount of rows
# print(sheet.max_row)
#
# # Retrieve the maximum amount of columns
# print(sheet.max_column)

# Press the green button in the gutter to run the script.
# if __name__ == '__main__':
#     main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
